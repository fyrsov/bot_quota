import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from bot.config import fmt_dt
from bot.database.models import ROLE_LABELS, Record, User
from bot.database.repositories.record_repo import RecordRepo
from bot.database.repositories.user_repo import UserRepo

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_BOLD = Font(bold=True)
_RETURN_FILL = PatternFill("solid", fgColor="FFD7D7")  # светло-красный для возвратов
_COL_WIDTHS = [5, 35, 15, 15, 20, 20, 20, 12]
_HEADERS = ["№", "ФИО", "Должность", "Telegram ID", "Номер договора", "Дата выдачи", "Дата возврата", "Статус"]


def _write_sheet(ws, records: list[Record], user_map: dict[int, User], title: str) -> None:
    ws.title = title[:31]

    ws.append(_HEADERS)
    for col_idx in range(1, len(_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    active_count = 0
    returned_count = 0
    for idx, record in enumerate(records, start=1):
        user = user_map.get(record.user_id)
        is_returned = record.is_cancelled
        returned_at = fmt_dt(record.cancelled_at) if is_returned else ""
        status = "Возврат" if is_returned else "Активна"
        row = [
            idx,
            user.full_name if user else f"ID:{record.user_id}",
            ROLE_LABELS.get(user.role, user.role) if user else "—",
            record.user_id,
            record.site_number,
            fmt_dt(record.created_at),
            returned_at,
            status,
        ]
        ws.append(row)
        if is_returned:
            returned_count += 1
            for col_idx in range(1, len(_HEADERS) + 1):
                ws.cell(row=ws.max_row, column=col_idx).fill = _RETURN_FILL
        else:
            active_count += 1

    ws.append([])
    ws.append(["", f"Активных: {active_count}  |  Возвратов: {returned_count}  |  Всего: {len(records)}", "", "", "", "", "", ""])
    total_row = ws.max_row
    ws.cell(row=total_row, column=2).font = _BOLD

    for col_idx, width in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width


async def build_excel(session: AsyncSession, months: list[str]) -> bytes:
    """
    Генерирует Excel-отчёт за список месяцев.
    Один месяц — один лист. Несколько — лист на каждый + сводный лист.
    months формат: ["2026-02", "2026-01", ...]
    """
    record_repo = RecordRepo(session)
    user_repo = UserRepo(session)

    users: list[User] = await user_repo.get_all()
    user_map: dict[int, User] = {u.telegram_id: u for u in users}

    wb = Workbook()
    wb.remove(wb.active)  # удаляем пустой лист по умолчанию

    all_records: list[Record] = []

    for month in sorted(months):
        records = await record_repo.get_by_month_all_users_full(month)
        all_records.extend(records)
        dt = datetime.strptime(month, "%Y-%m")
        sheet_title = dt.strftime("%B %Y")
        ws = wb.create_sheet(title=sheet_title)
        _write_sheet(ws, records, user_map, sheet_title)

    # Сводный лист если месяцев больше одного
    if len(months) > 1:
        ws_summary = wb.create_sheet(title="Сводная", index=0)
        _write_sheet(ws_summary, all_records, user_map, "Сводная")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
